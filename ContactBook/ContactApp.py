import time
import datetime
from tkinter import *
from tkinter import ttk, messagebox
import openpyxl as op
import mysql.connector
from PIL import Image, ImageTk
import webbrowser

try:
    db = mysql.connector.connect(host="localhost", user="root", password="root")
    cu = db.cursor()

    cu.execute("create database if not exists contactDb")
    cu.execute("use contactDb")
    cu.execute("create table if not exists numbers_(NAME varchar(30) not null, CONTACT varchar(10) not null,"
               " COMPANY varchar(10))")


    def select():
        cu.execute("select * from numbers_")
        return cu.fetchall()


    work = op.load_workbook("myXl.xlsx")
    data = work.active
    res = select()
    NAME = []
    if len(res) == 0:
        CONTACT = []
        COMPANY = []
        if data:
            for i in range(1, data.max_row):
                for j in data.iter_cols(1, 1):
                    NAME.append(j[i].value)

                for s in data.iter_cols(2, 2):
                    CONTACT.append(s[i].value)

                for n in data.iter_cols(3, 3):
                    COMPANY.append(n[i].value)
            for k in range(len(NAME)):
                cu.execute('insert into numbers_ values(%s, %s, %s)', (NAME[k], CONTACT[k], COMPANY[k]))
                db.commit()
            res = select()

        else:
            messagebox.showerror("File not Found")
    else:
        for i in range(1, data.max_row):
            for j in data.iter_cols(1, 1):
                NAME.append(j[i].value)


    def merge_sort(array):
        if len(array) > 1:
            l_half = array[:len(array) // 2:]
            r_half = array[len(array) // 2:]
            merge_sort(l_half)
            merge_sort(r_half)
            o = p = q = 0
            while o < len(l_half) and p < len(r_half):
                if l_half[o] <= r_half[p]:
                    array[q] = l_half[o]
                    o += 1
                else:
                    array[q] = r_half[p]
                    p += 1
                q += 1
            while o < len(l_half):
                array[q] = l_half[o]
                o += 1
                q += 1
            while p < len(r_half):
                array[q] = r_half[p]
                p += 1
                q += 1
        return array


    def bin_search(array, low, high, find):
        mid = (low + high) // 2
        if array[mid] == find:
            return mid
        elif low >= high:
            return -1
        elif array[mid] > find:
            return bin_search(array, low, mid - 1, find)
        elif array[mid] < find:
            return bin_search(array, mid + 1, high, find)


    NAME = merge_sort(NAME)


    def details_frame(name):

        cu.execute("use contactDb")
        cu.execute(f"select * from numbers_ where NAME='{name}'")
        result = cu.fetchall()

        cu.execute("use contact_history")
        cu.execute(f"create table if not exists s{result[0][1]}(DATETIME varchar(50))")

        def what_app():
            webbrowser.open("https://web.whatsapp.com/")

        def mails_():
            webbrowser.open("https://mail.google.com/")

        def exit_():
            detail.destroy()
            saved_frame()

        def make_call():
            cu.execute("use contact_history")
            cu.execute(f"insert into history values('{name}')")
            s_ = time.strftime("%H:%M:%S", time.localtime())
            v = datetime.date.today()
            cu.execute(f"insert into s{result[0][1]} values('{s_}       {str(v)}')")
            view.delete(0, END)
            cu.execute(f"select * from s{result[0][1]}")
            for r1 in cu:
                view.insert(END, r1[0])

        cu.execute("use contactDb")
        cu.execute(f"select * from numbers_ where NAME='{name}'")
        result = cu.fetchall()
        detail = Frame(win, height=540, width=400, bg="black")
        detail.place(x=0, y=61)

        back = Button(detail, image=BACK, bd=0, activebackground="black", bg="black", command=exit_)
        back.place(x=20, y=20)

        user = Label(detail, image=userIOC, bd=0, bg="black", activebackground="black")
        user.place(x=125, y=15)

        names = Label(detail, text=f"{result[0][0]}", font="Ariel 25 bold", bg="black", fg="white")
        names.place(x=20, y=185)

        calls_ = Button(detail, image=call, bd=0, bg="black", activebackground="black", command=make_call)
        calls_.place(x=333, y=185)

        numbers_ = Label(detail, text=f"+91 {result[0][1]}", font="Ariel 20 bold", bg="black", fg="white")
        numbers_.place(x=20, y=250)

        mobile = Label(detail, text=f"Mobile  |  {result[0][2]}", font="Ariel 10 bold", bg="black", fg="white")
        mobile.place(x=25, y=290)

        what_ = Button(detail, image=what, bd=0, bg="black", activebackground="black", command=what_app)
        what_.place(x=330, y=250)

        mail_vid = Label(detail, text="Video call or Mail", font="Ariel 15 bold", bg="black", fg="white")
        mail_vid.place(x=22, y=330)

        mails = Button(detail, image=mail, bd=0, bg="black", activebackground="black", command=mails_)
        mails.place(x=250, y=330)

        video_call = Button(detail, image=ved, bd=0, bg="black", activebackground="black", command=what_app)
        video_call.place(x=330, y=325)

        sep = ttk.Separator(detail, orient=HORIZONTAL)
        sep.place(relx=0, rely=0.70, relwidth=1, relheight=0.004)

        call_h = Label(detail, text="call history", font="georgia 13 bold", padx=148, bg="black", fg="white")
        call_h.place(x=0, y=382)

        view = Listbox(detail, font="ariel 15 bold", bd=0, height=5, width=36, justify=CENTER,
                       bg="#1a1a1a", fg="white", selectbackground="lightBlue")
        view.place(x=0, y=410)
        cu.execute("use contact_history")
        cu.execute(f"select * from s{result[0][1]}")
        for r in cu:
            view.insert(END, r[0])


    def saved_frame():
        def view_contact(event):
            name = box.get(box.curselection())
            saved.destroy()
            details_frame(name)

        def on_search():
            val = var.get()
            if val == '' or val == " ":
                messagebox.showerror("Empty", "Field is Empty")
            else:
                value = bin_search(NAME, 0, len(NAME) - 1, val)
                if value == -1:
                    messagebox.showerror("No Contact", "No Contact exists of this name")
                else:
                    d = messagebox.askyesno("Found", f"The Contact Found on {value + 1} number\n"
                                                     f"Do you Want to View contact details")
                    if d is True:
                        details_frame(val)

        saved = Frame(win, height=540, width=400, bg="#616362")
        saved.place(x=0, y=61)

        var = StringVar()
        en = Entry(saved, font="ariel 23", width=22, bg="#363535", bd=0, fg="white", textvariable=var)
        en.place(x=12, y=10)

        ser_bt = Button(saved, image=search, bd=0, activebackground="#363535", bg="#363535", command=on_search)
        ser_bt.place(x=344, y=10)
        box = Listbox(saved, bg="#0B0A0A", bd=0, height=13, fg="white", font="Georgia 23", width=21,
                      justify=CENTER)
        box.place(x=0, y=56)
        box.insert(END, *NAME)
        box.bind("<<ListboxSelect>>", view_contact)


    def call_his():
        def view_contact(event):
            name = box.get(box.curselection())
            his.destroy()
            details_frame(name)

        def search_():
            cu.execute("use contact_history")
            cu.execute(f"select * from history where name like '%{getter.get()}%'")
            box.delete(0, END)
            for s3 in cu:
                box.insert(END, s3[0])

        cu.execute("create database if not exists contact_history")
        cu.execute("use contact_history")
        cu.execute("create table if not exists history(NAME varchar(50))")
        his = Frame(win, height=540, width=400, bg="#787a79")
        his.place(x=0, y=61)

        getter = StringVar()
        en = Entry(his, font="ariel 23", width=22, bg="#363535", bd=0, fg="white", textvariable=getter)
        en.place(x=12, y=10)

        ser_bt = Button(his, image=search, bd=0, activebackground="#363535", bg="#363535", command=search_)
        ser_bt.place(x=344, y=10)
        box = Listbox(his, bg="#0d0d0d", bd=0, height=13, fg="white", font="Georgia 23", width=21,
                      justify=CENTER)
        box.place(x=0, y=56)
        box.delete(0, END)
        cu.execute("use contact_history")
        cu.execute("select * from history")
        for s2 in cu:
            box.insert(END, s2[0])
        box.bind("<<ListboxSelect>>", view_contact)


    win = Tk()
    win.wm_minsize(400, 600)
    win.wm_maxsize(400, 600)
    win.title("My Contact")

    call = ImageTk.PhotoImage(Image.open("Images/CALL.ico"))
    contact = ImageTk.PhotoImage(Image.open("Images/Contact.ico"))
    search = ImageTk.PhotoImage(Image.open("Images/SEARCH.ico"))

    userIOC = ImageTk.PhotoImage(Image.open("Images/user.ico"))
    mail = ImageTk.PhotoImage(Image.open("Images/MAIL.ico"))
    ved = ImageTk.PhotoImage(Image.open("Images/vid.ico"))
    what = ImageTk.PhotoImage(Image.open("Images/WHAT.ico"))
    BACK = ImageTk.PhotoImage(Image.open("Images/Back.ico"))

    can = Canvas(win, height=600, width=400, bg="#787a79", bd=0, relief=SOLID)
    can.pack()

    header = Frame(win, height=60, width=400, bg="black", bd=0)
    header.place(x=0, y=0)

    calls = Button(header, image=call, bg="black", bd=0, activebackground="black", command=call_his)
    calls.place(x=90, y=10)

    con = Button(header, image=contact, bg="black", bd=0, activebackground="black", command=saved_frame)
    con.place(x=270, y=10)

    call_his()
    win.mainloop()

except Exception as E:
    messagebox.showerror("Tkinter Error", f"Fatal Error {E}")
